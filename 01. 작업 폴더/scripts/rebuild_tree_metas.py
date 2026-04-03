# -*- coding: utf-8 -*-
"""학습맵__AI용.xlsx에서 tree.js + metas.js를 생성하는 스크립트"""
import sys, io, os, json, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

XLSX_PATH = os.path.join(os.path.dirname(__file__), '..', '..', '02. AI에게 제공한 자료들', '학습맵__AI용.xlsx')
OUT_TREE = os.path.join(os.path.dirname(__file__), '..', 'src', 'data', 'tree.js')
OUT_METAS = os.path.join(os.path.dirname(__file__), '..', 'src', 'data', 'metas.js')

# 학년 키 매핑
GRADE_MAP = {
    'E3': 'e3', 'E4': 'e4', 'E5': 'e5', 'E6': 'e6',
    'M7': 'm1', 'M8': 'm2', 'M9': 'm3', 'H10': 'h_cs1'
}
GRADE_META = {
    'e3': {'label': '초등 3학년', 'school': 'elementary', 'gradeNum': 3},
    'e4': {'label': '초등 4학년', 'school': 'elementary', 'gradeNum': 4},
    'e5': {'label': '초등 5학년', 'school': 'elementary', 'gradeNum': 5},
    'e6': {'label': '초등 6학년', 'school': 'elementary', 'gradeNum': 6},
    'm1': {'label': '중학교 1학년', 'school': 'middle', 'gradeNum': 1},
    'm2': {'label': '중학교 2학년', 'school': 'middle', 'gradeNum': 2},
    'm3': {'label': '중학교 3학년', 'school': 'middle', 'gradeNum': 3},
    'h_cs1': {'label': '고등 공통수학1', 'school': 'high', 'gradeNum': 1},
    'h_cs2': {'label': '고등 공통수학2', 'school': 'high', 'gradeNum': 2},
}

# 영역 추정 (표준체계 코드 기반)
def guess_area(std_str):
    if not std_str: return "수학 전반"
    s = std_str.split('|')[0] if '|' in std_str else std_str
    # MATA01~MATA03 = 수와 연산, MATB/MATA03 = 도형과 측정, MATC = 도형, MATD = 자료와 가능성, MATE = 규칙성
    if re.search(r'MATA0[1-2]', s): return "수와 연산"
    if re.search(r'MATA03|MATB', s): return "도형과 측정"
    if re.search(r'MATC', s): return "도형"
    if re.search(r'MATD|MATA04', s): return "자료와 가능성"
    if re.search(r'MATE|MATA05', s): return "규칙성"
    return "수학 전반"

# 노드 유형별 메타 기본값
def node_type_defaults(leaf_name, d2_name):
    name = leaf_name.strip()
    if 'AI 진단' in name or 'AI 진단' in d2_name:
        return {'ch': '진단', 'tp': '객관식(4지선다)', 'el': '지식이해', 'bl': '이해', 'df': '중', 'sc': 3, 'tm': 90}
    if 'AI 형성' in name or 'AI 형성' in d2_name:
        return {'ch': '총괄', 'tp': '객관식(4지선다)', 'el': '과정기능', 'bl': '적용', 'df': '중', 'sc': 3, 'tm': 90}
    if 'AI 익힘' in name:
        return {'ch': '연습', 'tp': '객관식(4지선다)', 'el': '과정기능', 'bl': '적용', 'df': '중', 'sc': 3, 'tm': 90}
    if '뚝딱' in name or '창의' in name or '곰곰' in name:
        return {'ch': '심화', 'tp': '서술형', 'el': '과정기능', 'bl': '분석', 'df': '상', 'sc': 5, 'tm': 180}
    if '스스로' in name or '자기 평가' in name or '단원 마무리' in name:
        return {'ch': '총괄', 'tp': '객관식(4지선다)', 'el': '과정기능', 'bl': '적용', 'df': '중', 'sc': 3, 'tm': 90}
    if '한눈에' in name or '개념 요약' in name or '개념 퀴즈' in name:
        return {'ch': '정리', 'tp': '객관식(4지선다)', 'el': '지식이해', 'bl': '기억', 'df': '하', 'sc': 2, 'tm': 60}
    if '함께 놀아요' in name or '수학 더하기' in name or '수학 놀이터' in name or '프로젝트' in name:
        return {'ch': '활동', 'tp': '서술형', 'el': '태도가치', 'bl': '창의', 'df': '중', 'sc': 4, 'tm': 120}
    if '알고 있나요' in name or '준비해요' in name or '준비하기' in name:
        return {'ch': '진단', 'tp': '객관식(4지선다)', 'el': '지식이해', 'bl': '기억', 'df': '하', 'sc': 2, 'tm': 60}
    if '대단원 도입' in name:
        return {'ch': '도입', 'tp': '객관식(4지선다)', 'el': '지식이해', 'bl': '이해', 'df': '하', 'sc': 2, 'tm': 60}
    if '해결해요' in name or '스스로 해결' in name:
        return {'ch': '연습', 'tp': '서술형', 'el': '과정기능', 'bl': '적용', 'df': '중', 'sc': 3, 'tm': 90}
    if '나의 미래' in name or '수학 실험실' in name:
        return {'ch': '활동', 'tp': '서술형', 'el': '태도가치', 'bl': '창의', 'df': '중', 'sc': 4, 'tm': 120}
    # 중등/고등 개념 노드 (기본값)
    return {'ch': '연습', 'tp': '객관식(4지선다)', 'el': '과정기능', 'bl': '이해', 'df': '중', 'sc': 3, 'tm': 90}

def make_key(grade_key, sem, d1_num, d2_num, d3_num, leaf_name):
    """고유 키 생성: 학년_학기_d1_d2_d3"""
    safe = re.sub(r'[^a-zA-Z0-9가-힣]', '', leaf_name)[:10]
    base = f"{grade_key}_s{sem}_{d1_num}_{d2_num}_{d3_num}"
    return base

def main():
    print("Loading Excel...")
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # 1단계: 엑셀 파싱
    raw_rows = []
    for r in range(3, ws.max_row + 1):
        school = ws.cell(r, 7).value
        grade = ws.cell(r, 8).value
        if not school or not grade: continue

        sem = str(ws.cell(r, 9).value or '0')
        d1_num = str(ws.cell(r, 11).value or '00').zfill(2)
        d1_name = (ws.cell(r, 12).value or '').strip()
        d2_num = str(ws.cell(r, 13).value or '00').zfill(2)
        d2_name = (ws.cell(r, 14).value or '').strip()
        d3_num = str(ws.cell(r, 15).value or '00').zfill(2)
        d3_name = (ws.cell(r, 16).value or '').strip()
        std = (ws.cell(r, 20).value or '').strip()

        gk = GRADE_MAP.get(f'{school}{grade}')
        if not gk: continue

        leaf_name = d3_name if d3_name else d2_name
        if not leaf_name: continue

        raw_rows.append({
            'gk': gk, 'sem': sem,
            'd1_num': d1_num, 'd1_name': d1_name,
            'd2_num': d2_num, 'd2_name': d2_name,
            'd3_num': d3_num, 'd3_name': d3_name,
            'leaf': leaf_name, 'std': std
        })

    print(f"Parsed {len(raw_rows)} rows")

    # 2단계: tree.js 구축 (학년 → 대단원 → 중단원 → 소단원)
    tree_by_grade = {}
    metas = {}
    key_counter = {}  # 동일 키 충돌 방지

    for row in raw_rows:
        gk = row['gk']
        if gk not in tree_by_grade:
            tree_by_grade[gk] = {}

        # 대단원 (D1)
        d1_key = f"s{row['sem']}_{row['d1_num']}"
        if d1_key not in tree_by_grade[gk]:
            tree_by_grade[gk][d1_key] = {
                'n': row['d1_num'], 'nm': row['d1_name'], 'sem': row['sem'],
                'children': {}
            }
        d1 = tree_by_grade[gk][d1_key]

        # 중단원 (D2)
        d2_key = row['d2_num']
        if d2_key not in d1['children']:
            d1['children'][d2_key] = {
                'n': row['d2_num'], 'nm': row['d2_name'],
                'children': []
            }
        d2 = d1['children'][d2_key]

        # 소단원/leaf (D3)
        leaf_key = make_key(gk, row['sem'], row['d1_num'], row['d2_num'], row['d3_num'], row['leaf'])

        # 키 충돌 방지
        if leaf_key in key_counter:
            key_counter[leaf_key] += 1
            leaf_key = f"{leaf_key}_{key_counter[leaf_key]}"
        else:
            key_counter[leaf_key] = 0

        leaf_node = {
            'n': row['d3_num'], 'nm': row['leaf'], 'k': leaf_key
        }
        if row['std']:
            leaf_node['std'] = row['std'].split('|')[:5]  # 최대 5개

        d2['children'].append(leaf_node)

        # metas 항목 생성
        defaults = node_type_defaults(row['leaf'], row['d2_name'])
        area = guess_area(row['std'])
        std_list = row['std'].split('|')[:5] if row['std'] else []

        metas[leaf_key] = {
            **defaults,
            'ar': area,
            'l1': row['d1_name'],
            'ac': std_list[0] if std_list else '',
            'ad': f"{row['d1_name']} > {row['d2_name']} > {row['leaf']}",
            'gd': f"{row['leaf']} 관련 문항을 출제하라. {row['d2_name']} 단원의 내용을 반영하여, 학생이 해당 개념을 이해하고 적용할 수 있는지 평가하는 문항이어야 한다.",
            'std': std_list,
            '_gradeKey': gk,
            'depth': {
                'd1': {'name': row['d1_name'], 'context': f"{row['d1_name']} 단원 전체.", 'w': 0.2},
                'd2': {'name': row['d2_name'] or row['d1_name'], 'context': f"{row['d2_name'] or row['d1_name']} 중단원.", 'w': 0.3},
                'd3': {'name': row['leaf'], 'context': f"{row['leaf']} 학습 활동.", 'w': 0.5}
            }
        }

    print(f"Generated {len(metas)} meta entries")

    # 3단계: tree.js 파일 생성
    print("Writing tree.js...")
    with open(OUT_TREE, 'w', encoding='utf-8') as f:
        f.write('export const GRADE_META = {\n')
        for gk, gm in sorted(GRADE_META.items()):
            f.write(f'  {gk}: {{ label: "{gm["label"]}", school: "{gm["school"]}", gradeNum: {gm["gradeNum"]} }},\n')
        f.write('};\n\n')
        f.write('export const TREE_BY_GRADE = {\n')

        for gk in ['e3', 'e4', 'e5', 'e6', 'm1', 'm2', 'm3', 'h_cs1']:
            if gk not in tree_by_grade:
                f.write(f'\n  /* ── {GRADE_META.get(gk, {}).get("label", gk)} ── */\n')
                f.write(f'  {gk}: [],\n')
                continue

            label = GRADE_META.get(gk, {}).get('label', gk)
            f.write(f'\n  /* ── {label} {"─" * max(1, 50 - len(label) * 2)} */\n')
            f.write(f'  {gk}: [\n')

            d1_items = sorted(tree_by_grade[gk].items())
            for d1_key, d1 in d1_items:
                f.write(f'    {{ n:"{d1["n"]}", nm:"{d1["nm"]}", ch:[\n')

                d2_items = sorted(d1['children'].items())
                for d2_key, d2 in d2_items:
                    d2_name_esc = d2['nm'].replace('"', '\\"')
                    if len(d2['children']) == 0:
                        f.write(f'      {{ n:"{d2["n"]}", nm:"{d2_name_esc}", ch:[] }},\n')
                    else:
                        f.write(f'      {{ n:"{d2["n"]}", nm:"{d2_name_esc}", ch:[\n')
                        for leaf in d2['children']:
                            nm_esc = leaf['nm'].replace('"', '\\"').strip()
                            std_str = ''
                            if leaf.get('std'):
                                std_str = f', std:{json.dumps(leaf["std"])}'
                            f.write(f'        {{n:"{leaf["n"]}",nm:"{nm_esc}",k:"{leaf["k"]}"{std_str}}},\n')
                        f.write(f'      ]}},\n')

                f.write(f'    ]}},\n')

            f.write(f'  ],\n')

        # h_cs2는 학습맵에 없으므로 빈 배열
        f.write(f'\n  /* ── 고등 공통수학2 ── (학습맵 미제공) */\n')
        f.write(f'  h_cs2: [],\n')
        f.write('};\n')

    tree_size = os.path.getsize(OUT_TREE)
    print(f"tree.js: {tree_size:,} bytes")

    # 4단계: metas.js 파일 생성
    print("Writing metas.js...")
    with open(OUT_METAS, 'w', encoding='utf-8') as f:
        f.write('export const METAS = {\n')

        current_grade = None
        for key in sorted(metas.keys()):
            m = metas[key]
            gk = m['_gradeKey']

            if gk != current_grade:
                current_grade = gk
                label = GRADE_META.get(gk, {}).get('label', gk)
                f.write(f'\n  /* ── {label} {"─" * max(1, 50 - len(label) * 2)} */\n')

            std_json = json.dumps(m['std'], ensure_ascii=False)
            gd_esc = m['gd'].replace('"', '\\"').replace('\n', ' ')
            ad_esc = m['ad'].replace('"', '\\"')
            d1_ctx = m['depth']['d1']['context'].replace('"', '\\"')
            d2_ctx = m['depth']['d2']['context'].replace('"', '\\"')
            d3_ctx = m['depth']['d3']['context'].replace('"', '\\"')
            d1_nm = m['depth']['d1']['name'].replace('"', '\\"')
            d2_nm = m['depth']['d2']['name'].replace('"', '\\"')
            d3_nm = m['depth']['d3']['name'].replace('"', '\\"')

            f.write(f'  "{key}":{{ch:"{m["ch"]}",tp:"{m["tp"]}",el:"{m["el"]}",bl:"{m["bl"]}",df:"{m["df"]}",sc:{m["sc"]},tm:{m["tm"]},')
            f.write(f'ar:"{m["ar"]}",l1:"{m["l1"]}",ac:"{m["ac"]}",ad:"{ad_esc}",')
            f.write(f'gd:"{gd_esc}",')
            f.write(f'std:{std_json},_gradeKey:"{gk}",')
            f.write(f'depth:{{d1:{{name:"{d1_nm}",context:"{d1_ctx}",w:0.2}},d2:{{name:"{d2_nm}",context:"{d2_ctx}",w:0.3}},d3:{{name:"{d3_nm}",context:"{d3_ctx}",w:0.5}}}}}},\n')

        f.write('};\n')

    metas_size = os.path.getsize(OUT_METAS)
    print(f"metas.js: {metas_size:,} bytes")

    # 통계
    grade_counts = {}
    for key, m in metas.items():
        gk = m['_gradeKey']
        if gk not in grade_counts: grade_counts[gk] = 0
        grade_counts[gk] += 1

    print(f"\n=== 생성 완료 ===")
    print(f"전체 leaf 노드: {len(metas)}개")
    for gk in ['e3', 'e4', 'e5', 'e6', 'm1', 'm2', 'm3', 'h_cs1']:
        cnt = grade_counts.get(gk, 0)
        print(f"  {gk}: {cnt}개")

if __name__ == '__main__':
    main()
